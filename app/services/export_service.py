"""
导出服务。

Excel 导出使用 `docs/产销预测及供应表.xlsx` 的第二个 Sheet 作为汇总模板。
"""

from __future__ import annotations

from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill


PROJECT_ROOT = Path(__file__).resolve().parents[2]
DEFAULT_TEMPLATE_PATH = PROJECT_ROOT / "docs" / "产销预测及供应表.xlsx"


class ExportService:
    """导出服务"""

    DETAIL_HEADERS = [
        "事业部",
        "产品系列",
        "生产工厂",
        "属性",
        "型号",
        "产品描述",
        "产品类别",
        "出厂价",
        "上市日期",
        "4.30日库存",
        "5月排产（4.30-5.29）",
        "5月总供应",
        "省大区SAR",
        "网络经销商SAR",
        "电商直营SAR",
        "KA部SAR",
        "拓展部SAR",
        "5月SAR合计",
        "5月SAR差异",
        "可满足合计",
        "省大区（可满足）",
        "网络经销商（可满足）",
        "电商直营（可满足）",
        "KA部（可满足）",
        "拓展部（可满足）",
        "KA-5月25日前可满足（含25日）",
        "未满足合计",
        "省大区（未满足）",
        "网络经销商（未满足）",
        "电商直营（未满足）",
        "KA部（未满足）",
        "拓展部（未满足）",
        "风险等级",
        "风险评分",
        "风险原因",
    ]

    def generate_export_filename(self, export_type: str, export_format: str, task_id: int | None = None) -> str:
        """生成导出文件名。"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        task_suffix = f"_task{task_id}" if task_id else ""
        return f"export_{export_type.lower()}{task_suffix}_{timestamp}.{export_format.lower()}"

    def export_to_excel(
        self,
        data: list[dict[str, Any]],
        output_path: str,
        sheet_name: str = "SKU明细",
        include_summary: bool = True,
        summary_data: dict[str, Any] | None = None,
    ) -> str:
        """
        兼容旧调用的 Excel 导出入口。

        新实现统一使用第二个 Sheet 汇总模板；`sheet_name/include_summary/summary_data`
        保留为兼容参数。
        """
        return self.export_summary_template(data, output_path)

    def export_summary_template(
        self,
        sku_rows: list[dict[str, Any]],
        output_path: str,
        template_path: Path = DEFAULT_TEMPLATE_PATH,
    ) -> str:
        """基于第二个 Sheet 汇总模板导出 Excel。"""
        wb = load_workbook(template_path)
        template_ws = wb.worksheets[1]
        template_ws.title = "预测及排期汇总表"

        # 清空模板数据区，保留表头、合并单元格、样式和公式区域。
        for row in template_ws.iter_rows(min_row=3, max_row=template_ws.max_row, max_col=20):
            for cell in row:
                if not isinstance(cell, MergedCell):
                    cell.value = None

        grouped = self._group_by_series(sku_rows)
        row_idx = 3
        for series_name, values in grouped.items():
            template_ws.cell(row_idx, 1, "按项目呈现")
            template_ws.cell(row_idx, 2, series_name or "未分类")
            for col_idx, value in enumerate(values, start=3):
                template_ws.cell(row_idx, col_idx, value)
            row_idx += 1

        if grouped:
            template_ws.cell(row_idx, 1, "按项目呈现")
            template_ws.cell(row_idx, 2, "合计")
            for col_idx in range(3, 21):
                letter = get_column_letter(col_idx)
                template_ws.cell(row_idx, col_idx, f"=SUM({letter}3:{letter}{row_idx - 1})")
            for cell in template_ws[row_idx]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")

        detail_ws = wb.create_sheet("SKU明细")
        for col_idx, header in enumerate(self.DETAIL_HEADERS, start=1):
            cell = detail_ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(color="FFFFFF", bold=True)
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        for row_idx, row in enumerate(sku_rows, start=2):
            for col_idx, header in enumerate(self.DETAIL_HEADERS, start=1):
                detail_ws.cell(row=row_idx, column=col_idx, value=row.get(header))

        for ws in [template_ws, detail_ws]:
            for column_cells in ws.columns:
                letter = get_column_letter(column_cells[0].column)
                max_len = max(len(str(cell.value or "")) for cell in column_cells[:200])
                ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 38)

        wb.save(output_path)
        return output_path

    def export_to_csv(self, data: list[dict[str, Any]], output_path: str, encoding: str = "utf-8-sig") -> str:
        """导出 CSV。"""
        if not data:
            raise ValueError("数据为空，无法导出")
        pd.DataFrame(data).to_csv(output_path, index=False, encoding=encoding)
        return output_path

    def _group_by_series(self, sku_rows: list[dict[str, Any]]) -> dict[str, list[float]]:
        """按产品系列聚合到第二个 Sheet 的 18 个指标列。"""
        buckets: dict[str, list[float]] = defaultdict(lambda: [0.0] * 18)
        mapping = [
            ("省大区SAR", 0),
            ("省大区（可满足）", 1),
            ("省大区（未满足）", 2),
            ("网络经销商SAR", 3),
            ("网络经销商（可满足）", 4),
            ("网络经销商（未满足）", 5),
            ("电商直营SAR", 6),
            ("电商直营（可满足）", 7),
            ("电商直营（未满足）", 8),
            ("KA部SAR", 9),
            ("KA部（可满足）", 10),
            ("KA部（未满足）", 11),
            ("拓展部SAR", 12),
            ("拓展部（可满足）", 13),
            ("拓展部（未满足）", 14),
            ("5月SAR合计", 15),
            ("可满足合计", 16),
            ("未满足合计", 17),
        ]
        for row in sku_rows:
            series = str(row.get("产品系列") or "未分类")
            for key, index in mapping:
                buckets[series][index] += float(row.get(key) or 0)
        return dict(buckets)
