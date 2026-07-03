"""
API 端到端测试 - 使用 FastAPI TestClient
"""

import pytest
from fastapi.testclient import TestClient
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
import sys
from pathlib import Path
from decimal import Decimal

# 添加项目根目录到路径
sys.path.insert(0, str(Path(__file__).parent.parent))

from main import app
from app.db import get_db
from app.models.base import Base
from app.models.task import TaskFile, TaskSummary
from app.models.sku import SkuForecastDetail
from app.models.user import User
from app.models.ai import AIConversation
from app.models.audit import AuditLog, ExportRecord

# 创建测试数据库（内存 SQLite）
SQLALCHEMY_DATABASE_URL = "sqlite:///./test.db"
engine = create_engine(SQLALCHEMY_DATABASE_URL, connect_args={"check_same_thread": False})
TestingSessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)

# 创建测试表。先 drop 避免仓库根目录旧 test.db 遗留历史字段结构。
Base.metadata.drop_all(bind=engine)
Base.metadata.create_all(bind=engine)


def override_get_db():
    """覆盖数据库依赖"""
    db = TestingSessionLocal()
    try:
        yield db
    finally:
        db.close()


app.dependency_overrides[get_db] = override_get_db
client = TestClient(app)


class TestAPIEndpoints:
    """API 端点测试"""

    def _create_success_task(self, tmp_path, task_id=1001):
        """创建带汇总和 SKU 明细的成功任务。"""
        db = TestingSessionLocal()
        try:
            db.query(AIConversation).delete()
            db.query(AuditLog).delete()
            db.query(ExportRecord).delete()
            db.query(SkuForecastDetail).delete()
            db.query(TaskSummary).delete()
            db.query(TaskFile).delete()
            db.query(User).delete()
            db.commit()
            db.add(User(id=1, username="planner", password_hash="x", real_name="陈计划", role="PLANNER"))
            db.commit()

            task = TaskFile(
                id=task_id,
                task_name="测试任务",
                file_name="forecast.xlsx",
                file_path=str(tmp_path / "forecast.xlsx"),
                file_size=100,
                file_md5="abc",
                upload_user_id=1,
                status="SUCCESS",
                progress=100,
            )
            db.add(task)
            db.commit()
            db.refresh(task)

            db.add(
                TaskSummary(
                    id=task_id + 1000,
                    task_id=task.id,
                    month="2024-05",
                    total_supply=Decimal("1000"),
                    total_sar=Decimal("1500"),
                    total_gap=Decimal("-500"),
                    service_level=Decimal("66.67"),
                    target_service_level=Decimal("98"),
                    critical_risk_count=1,
                    high_risk_count=0,
                    medium_risk_count=0,
                )
            )
            db.add(
                SkuForecastDetail(
                    id=task_id + 2000,
                    task_id=task.id,
                    month="2024-05",
                    product_name="无线耳机 B",
                    sku_code="SKU-002",
                    material_code="MAT-002",
                    business_unit="儿童",
                    product_series="量子战队3",
                    factory="ODM",
                    product_attribute="新品",
                    product_category="常规生产产品",
                    ex_factory_price=Decimal("49.5"),
                    initial_inventory=Decimal("100"),
                    production_plan=Decimal("100"),
                    safety_stock=Decimal("50"),
                    total_supply=Decimal("150"),
                    sar_province=Decimal("500"),
                    sar_dealer=Decimal("400"),
                    sar_ecommerce=Decimal("200"),
                    sar_ka=Decimal("100"),
                    sar_expansion=Decimal("0"),
                    sar_total=Decimal("1200"),
                    gap=Decimal("-1050"),
                    service_level=Decimal("12.5"),
                    satisfied_province=Decimal("62.5"),
                    satisfied_dealer=Decimal("50"),
                    satisfied_ecommerce=Decimal("25"),
                    satisfied_ka=Decimal("12.5"),
                    satisfied_expansion=Decimal("0"),
                    satisfied_demand=Decimal("150"),
                    satisfied_ka_before_25=Decimal("10"),
                    unsatisfied_province=Decimal("437.5"),
                    unsatisfied_dealer=Decimal("350"),
                    unsatisfied_ecommerce=Decimal("175"),
                    unsatisfied_ka=Decimal("87.5"),
                    unsatisfied_expansion=Decimal("0"),
                    unsatisfied_demand=Decimal("1050"),
                    risk_level="CRITICAL",
                    risk_reason="供应严重不足",
                    risk_score=Decimal("180"),
                )
            )
            db.commit()
            return task.id
        finally:
            db.close()

    def test_root(self):
        """测试根路径"""
        response = client.get("/")
        assert response.status_code == 200
        data = response.json()
        assert "app" in data
        assert "version" in data
        assert data["version"] == "2.2.0"
        print("PASS: 根路径测试")

    def test_health_check(self):
        """测试健康检查"""
        response = client.get("/api/health")
        assert response.status_code == 200
        data = response.json()
        assert "status" in data
        print("PASS: 健康检查测试")

    def test_login_invalid_credentials(self):
        """测试无效登录"""
        response = client.post(
            "/api/auth/login",
            json={"username": "invalid", "password": "wrong"}
        )
        # 用户不存在或密码错误都应该返回401
        # 如果数据库为空，可能返回401或422
        assert response.status_code in [401, 422]
        print("PASS: 无效登录被拒绝")

    def test_login_success_returns_token_and_audit_log(self):
        """成功登录应返回 token 并写入审计日志。"""
        import bcrypt

        db = TestingSessionLocal()
        try:
            db.query(AuditLog).delete()
            db.query(User).delete()
            db.commit()
            password_hash = bcrypt.hashpw(b"admin123", bcrypt.gensalt()).decode("utf-8")
            db.add(User(id=1, username="admin", password_hash=password_hash, real_name="系统管理员", role="ADMIN", status="ACTIVE"))
            db.commit()
        finally:
            db.close()

        response = client.post("/api/auth/login", json={"username": "admin", "password": "admin123"})
        assert response.status_code == 200
        data = response.json()
        assert data["success"] is True
        assert data["access_token"]

        db = TestingSessionLocal()
        try:
            assert db.query(AuditLog).filter(AuditLog.operation == "LOGIN").count() == 1
        finally:
            db.close()

    def test_logout(self):
        """测试登出"""
        response = client.post("/api/auth/logout")
        assert response.status_code == 200
        data = response.json()
        assert data["success"] is True
        print("PASS: 登出测试")

    def test_get_tasks_empty(self):
        """测试获取空任务列表"""
        response = client.get("/api/tasks")
        assert response.status_code == 200
        data = response.json()
        assert isinstance(data, list)
        print("PASS: 任务列表测试")

    def test_get_task_not_found(self):
        """测试获取不存在的任务"""
        response = client.get("/api/tasks/9999")
        assert response.status_code == 404
        print("PASS: 任务不存在返回404")

    def test_upload_invalid_file(self):
        """测试上传无效文件"""
        response = client.post(
            "/api/tasks/upload",
            files={"file": ("test.txt", b"test content", "text/plain")}
        )
        assert response.status_code == 400
        print("PASS: 无效文件被拒绝")

    def test_chat_context_includes_sku_details(self, monkeypatch, tmp_path):
        """AI 对话上下文应包含任务汇总和 SKU 明细"""
        captured = {}

        def fake_chat(self, session_id, user_message, context=None):
            captured["session_id"] = session_id
            captured["user_message"] = user_message
            captured["context"] = context
            return {
                "success": True,
                "message": "ok",
                "session_id": session_id,
                "type": "llm_generate",
                "model": "test-model",
                "timestamp": "2026-06-21T00:00:00",
            }

        monkeypatch.setattr("app.services.ai_service.AIService.chat", fake_chat)

        db = TestingSessionLocal()
        try:
            db.query(SkuForecastDetail).delete()
            db.query(TaskSummary).delete()
            db.query(TaskFile).delete()
            db.query(User).delete()
            db.commit()
            db.add(User(id=1, username="planner", password_hash="x", real_name="陈计划", role="PLANNER"))
            db.commit()

            task = TaskFile(
                id=1001,
                task_name="测试任务",
                file_name="forecast.xlsx",
                file_path=str(tmp_path / "forecast.xlsx"),
                file_size=100,
                file_md5="abc",
                upload_user_id=1,
                status="SUCCESS",
                progress=100,
            )
            db.add(task)
            db.commit()
            db.refresh(task)
            task_id = task.id

            db.add(
                TaskSummary(
                    id=2001,
                    task_id=task.id,
                    month="2024-05",
                    total_supply=Decimal("1000"),
                    total_sar=Decimal("1500"),
                    total_gap=Decimal("-500"),
                    service_level=Decimal("66.67"),
                    target_service_level=Decimal("98"),
                    critical_risk_count=1,
                    high_risk_count=0,
                    medium_risk_count=0,
                )
            )
            db.add(
                SkuForecastDetail(
                    id=3001,
                    task_id=task.id,
                    month="2024-05",
                    product_name="无线耳机 B",
                    sku_code="SKU-002",
                    material_code="MAT-002",
                    initial_inventory=Decimal("100"),
                    production_plan=Decimal("100"),
                    safety_stock=Decimal("50"),
                    total_supply=Decimal("150"),
                    sar_province=Decimal("500"),
                    sar_dealer=Decimal("400"),
                    sar_ecommerce=Decimal("200"),
                    sar_ka=Decimal("100"),
                    sar_expansion=Decimal("0"),
                    sar_total=Decimal("1200"),
                    gap=Decimal("-1050"),
                    service_level=Decimal("12.5"),
                    satisfied_province=Decimal("62.5"),
                    satisfied_dealer=Decimal("50"),
                    satisfied_ecommerce=Decimal("25"),
                    satisfied_ka=Decimal("12.5"),
                    satisfied_expansion=Decimal("0"),
                    satisfied_demand=Decimal("150"),
                    unsatisfied_province=Decimal("437.5"),
                    unsatisfied_dealer=Decimal("350"),
                    unsatisfied_ecommerce=Decimal("175"),
                    unsatisfied_ka=Decimal("87.5"),
                    unsatisfied_expansion=Decimal("0"),
                    unsatisfied_demand=Decimal("1050"),
                    risk_level="CRITICAL",
                    risk_reason="供应严重不足",
                    risk_score=Decimal("180"),
                )
            )
            db.commit()
        finally:
            db.close()

        response = client.post(
            "/api/chat",
            json={"session_id": "test-session", "message": "SKU-002 为什么缺货？", "task_id": task_id},
        )

        assert response.status_code == 200
        assert captured["context"]["task_summary"]["total_gap"] == -500.0
        assert captured["context"]["sku_details"][0]["sku_code"] == "SKU-002"
        assert captured["context"]["sku_details"][0]["risk_level"] == "CRITICAL"
        assert captured["context"]["sku_details"][0]["gap"] == -1050.0

    def test_chat_persists_and_lists_conversation_history(self, monkeypatch, tmp_path):
        """AI 对话应入库，并能按会话列表和详情查询。"""
        task_id = self._create_success_task(tmp_path, task_id=4001)

        def fake_chat(self, session_id, user_message, context=None):
            return {
                "success": True,
                "message": "第一轮回答",
                "session_id": session_id,
                "type": "llm_generate",
                "model": "test-model",
                "timestamp": "2026-06-21T00:00:00",
            }

        monkeypatch.setattr("app.services.ai_service.AIService.chat", fake_chat)

        response = client.post(
            "/api/chat",
            json={"session_id": "history-session", "message": "第一轮问题", "task_id": task_id},
        )

        assert response.status_code == 200
        data = response.json()
        assert data["success"] is True
        assert data["conversation_id"] > 0

        db = TestingSessionLocal()
        try:
            rows = db.query(AIConversation).filter(AIConversation.session_id == "history-session").all()
            assert len(rows) == 1
            assert rows[0].question == "第一轮问题"
            assert rows[0].answer == "第一轮回答"
            assert rows[0].task_id == task_id
        finally:
            db.close()

        list_response = client.get(f"/api/chat/sessions?task_id={task_id}")
        assert list_response.status_code == 200
        sessions = list_response.json()
        assert sessions[0]["session_id"] == "history-session"
        assert sessions[0]["title"] == "第一轮问题"
        assert sessions[0]["message_count"] == 1

        detail_response = client.get("/api/chat/sessions/history-session")
        assert detail_response.status_code == 200
        detail = detail_response.json()
        assert detail["session_id"] == "history-session"
        assert detail["messages"][0]["question"] == "第一轮问题"
        assert detail["messages"][0]["answer"] == "第一轮回答"

    def test_chat_uses_persisted_history_as_model_context(self, monkeypatch, tmp_path):
        """继续同一会话时，应把历史问答传入模型上下文。"""
        task_id = self._create_success_task(tmp_path, task_id=5001)
        captured = {}

        db = TestingSessionLocal()
        try:
            db.add(
                AIConversation(
                    id=1,
                    task_id=task_id,
                    user_id=1,
                    session_id="continue-session",
                    question="第一轮问题",
                    answer="第一轮回答",
                    context={"task_summary": {"total_gap": -500.0}},
                )
            )
            db.commit()
        finally:
            db.close()

        def fake_chat(self, session_id, user_message, context=None):
            captured["context"] = context
            return {
                "success": True,
                "message": "第二轮回答",
                "session_id": session_id,
                "type": "llm_generate",
                "model": "test-model",
                "timestamp": "2026-06-21T00:00:00",
            }

        monkeypatch.setattr("app.services.ai_service.AIService.chat", fake_chat)

        response = client.post(
            "/api/chat",
            json={"session_id": "continue-session", "message": "第二轮问题", "task_id": task_id},
        )

        assert response.status_code == 200
        assert captured["context"]["conversation_history"] == [
            {"role": "user", "content": "第一轮问题"},
            {"role": "assistant", "content": "第一轮回答"},
        ]

        detail_response = client.get("/api/chat/sessions/continue-session")
        assert detail_response.status_code == 200
        assert len(detail_response.json()["messages"]) == 2

    def test_chat_includes_long_term_memory_context(self, monkeypatch, tmp_path):
        """AI 对话应检索长期记忆，并在回答后继续写入项目内记忆。"""
        from app.core.config import get_settings
        from app.services.memory_service import AssistantMemoryService

        task_id = self._create_success_task(tmp_path, task_id=7001)
        memory_dir = tmp_path / "mem0"
        captured = {}

        monkeypatch.setenv("MEMORY_ENABLED", "true")
        monkeypatch.setenv("MEMORY_BACKEND", "jsonl")
        monkeypatch.setenv("MEMORY_DIR", str(memory_dir))
        get_settings.cache_clear()

        memory_service = AssistantMemoryService()
        memory_service.add_interaction(
            user_message="请记住：我最关注 KA 部缺口",
            assistant_message="已记录。",
            user_id=1,
            session_id="old-session",
        )

        def fake_chat(self, session_id, user_message, context=None):
            captured["context"] = context
            return {
                "success": True,
                "message": "长期记忆回答",
                "session_id": session_id,
                "type": "llm_generate",
                "model": "test-model",
                "timestamp": "2026-07-03T00:00:00",
            }

        monkeypatch.setattr("app.services.ai_service.AIService.chat", fake_chat)

        response = client.post(
            "/api/chat",
            json={"session_id": "memory-session", "message": "KA部缺口怎么处理？", "task_id": task_id},
        )

        assert response.status_code == 200
        assert "long_term_memories" in captured["context"]
        assert "KA 部缺口" in captured["context"]["long_term_memory_text"]
        assert (memory_dir / "memories.jsonl").exists()
        assert len((memory_dir / "memories.jsonl").read_text(encoding="utf-8").splitlines()) >= 2
        get_settings.cache_clear()

    def test_export_logs_settings_notifications_and_search(self, tmp_path):
        """导出、操作日志、设置、通知和搜索 API 应可用。"""
        from openpyxl import load_workbook
        from app.models.config import SystemConfig

        task_id = self._create_success_task(tmp_path, task_id=6001)
        db = TestingSessionLocal()
        try:
            db.merge(SystemConfig(config_key="file.max_size", config_value="52428800", config_type="NUMBER", description="文件上传大小上限"))
            db.commit()
        finally:
            db.close()

        export_response = client.post(f"/api/tasks/{task_id}/export?export_type=SKU_DETAIL&export_format=XLSX")
        assert export_response.status_code == 200
        assert export_response.headers["content-type"].startswith("application/vnd.openxmlformats")

        export_file = tmp_path / "export.xlsx"
        export_file.write_bytes(export_response.content)
        wb = load_workbook(export_file, data_only=False)
        assert "预测及排期汇总表" in wb.sheetnames
        assert "SKU明细" in wb.sheetnames
        assert wb["预测及排期汇总表"]["A1"].value == "呈现类别"
        assert wb["SKU明细"]["A1"].value == "事业部"

        logs_response = client.get("/api/logs")
        assert logs_response.status_code == 200
        operations = {item["operation"] for item in logs_response.json()["items"]}
        assert "EXPORT" in operations

        settings_response = client.get("/api/settings")
        assert settings_response.status_code == 200
        assert any(item["key"] == "file.max_size" for item in settings_response.json()["items"])

        notifications_response = client.get("/api/notifications")
        assert notifications_response.status_code == 200
        assert notifications_response.json()["items"]

        search_response = client.get("/api/search?q=SKU-002")
        assert search_response.status_code == 200
        assert search_response.json()["skus"][0]["title"] == "SKU-002"


if __name__ == "__main__":
    pytest.main([__file__, "-v", "-s"])
